import openpyxl
import json
import os
import glob
import re

WATER_LEVEL_DIR = r"C:\Users\letsm\Downloads\water level"
OUTPUT_DIR = r"C:\Users\letsm\OneDrive\Desktop\JalDrishti\Datasets\waterLevelJson"


def slugify(name: str) -> str:
    name = re.sub(r"[^\w\s-]", "", name).strip()
    return re.sub(r"[\s]+", "_", name)


def extract_station_metadata(ws) -> dict:
    station = {}
    reading = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Station Data":
            reading = True
            continue
        if reading and row[0] and row[1]:
            station[str(row[0]).strip()] = str(row[1]).strip()
    return station


def extract_readings(ws) -> list:
    rows = list(ws.iter_rows(values_only=True))
    try:
        header_idx = next(i for i, r in enumerate(rows) if r[0] == "Data Type Code")
    except StopIteration:
        return []
    readings = []
    for r in rows[header_idx + 1 :]:
        if r[0] != "HGZ":
            continue
        readings.append(
            {
                "data_type_code": r[0],
                "data_type_description": r[1],
                "timestamp": r[2].isoformat() if hasattr(r[2], "isoformat") else str(r[2]),
                "water_level_m": round(float(r[3]), 4) if r[3] is not None else None,
                "unit": r[4],
            }
        )
    return readings


def convert_xlsx(path: str, out_dir: str):
    wb = openpyxl.load_workbook(path)

    metadata_sheet = next((s for s in wb.sheetnames if s.lower().startswith("metadata")), None)
    readings_sheet = next((s for s in wb.sheetnames if "ground water level" in s.lower() and not s.lower().startswith("metadata")), None)

    station = extract_station_metadata(wb[metadata_sheet]) if metadata_sheet else {}
    readings = extract_readings(wb[readings_sheet]) if readings_sheet else []

    output = {"station": station, "readings": readings}

    base = os.path.splitext(os.path.basename(path))[0]
    out_name = slugify(base) + ".json"
    out_path = os.path.join(out_dir, out_name)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)

    return out_name, len(readings)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    files = glob.glob(os.path.join(WATER_LEVEL_DIR, "*.xlsx"))
    print(f"Found {len(files)} xlsx files\n")

    for path in sorted(files):
        try:
            name, count = convert_xlsx(path, OUTPUT_DIR)
            print(f"  OK  {name}  ({count} readings)")
        except Exception as e:
            print(f"  ERR {os.path.basename(path)}: {e}")

    print(f"\nDone. JSON files saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
