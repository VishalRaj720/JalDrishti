"""Row-level control over the files the physics engine actually reads.

WHY THIS EXISTS. `dataset_sync.py` carries approved field observations *into*
`Datasets/`. It is append-only, so until now a row that landed there was
permanent: a typo in a field officer's grade estimate, or a sighting later shown
to be something else, could be approved into the engine's inputs and never taken
back out. This module is the other half — list, edit, delete, restore — and the
provenance column that makes any of it safe.

THE `record_source` COLUMN. Every writable dataset gains one, values `original`
or `added`, backfilled to `original` the first time the file is touched.

    original   shipped with the project — CGWB, UDEPO, GSI, NAQUIM
    added      written by this system from an approved field observation

**`original` rows are immutable.** Not editable, not deletable, enforced here in
the service layer with a 409 so every caller hits it rather than in the UI where
only a browser would. A screening tool whose operator can quietly rewrite its own
evidence base cannot defend a number, and the whole point of the column is that
the distinction is mechanical rather than a matter of remembering.

WHY NOT JUST `source`. The obvious name was taken, twice, meaning something else:
`cgwb_waterlevel_jharkhand.csv` has `source` = the collecting agency (`CGWB`),
and `naquim_reference/naquim_vertical.csv` has `source` = the citation a row was
extracted from. Overloading either would destroy real data and mislead anyone
reading the file directly, so the provenance marker gets its own unambiguous
name and uses it in every file.

ROW IDENTITY is the file's own natural key, never a positional index — indices
shift under deletion, and an admin acting on a stale list would hit the wrong
row. Each registry entry names its id column.

EVERY MUTATION: back up first, write, clear the pipeline's caches, audit. The
backup is what makes this reversible, and `restore()` is the way back.
"""
from __future__ import annotations

import io
import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from loguru import logger

from app.exceptions import AppException

REPO_ROOT = Path(__file__).resolve().parents[3]
DATASETS = REPO_ROOT / "Datasets"

#: The provenance column. See the module docstring for why it is not `source`.
SOURCE_COL = "record_source"
SOURCE_ORIGINAL = "original"
SOURCE_ADDED = "added"

#: The sync batch that wrote an `added` row; empty on `original` rows.
#:
#: Without it, deleting a row could not free the observation that produced it —
#: nothing else links a line in a CSV back to a `field_observations` id — and the
#: portal would go on reporting an observation as "in the model" after its row had
#: been removed from the file the model reads. Encoding the link in a free-text
#: `notes` field was tried first and is not sound: two of these files have no
#: notes column, and a human editing notes would silently break provenance.
REF_COL = "record_ref"


class DatasetError(AppException):
    def __init__(self, message: str, status_code: int = 409):
        super().__init__(message, status_code=status_code)


@dataclass(frozen=True)
class DatasetFile:
    """One file the engine reads and this system may write."""

    key: str            #: URL-safe identifier used by the API
    relpath: str        #: path under Datasets/
    kind: str           #: "csv" | "xlsx"
    id_column: str      #: the file's own natural key
    label: str
    governs: str        #: what changes in the model when this file changes
    stale_marks: tuple[str, ...] = ()   #: derived artifacts invalidated by a write
    header_row: int = 0                 #: xlsx only — rows of preamble above the header

    @property
    def path(self) -> Path:
        return DATASETS / self.relpath


#: Everything the application may write. Reference geography (district, aquifer
#: and river GeoJSON) is deliberately absent: it is replaced wholesale through
#: `/ingest/*`, which checksums the source file and writes a `dataset_versions`
#: row, not edited row by row.
REGISTRY: dict[str, DatasetFile] = {
    "water_quality": DatasetFile(
        key="water_quality",
        relpath="waterQuality_jharkhand.csv",
        kind="csv",
        id_column="S. No.",
        label="Groundwater chemistry (CGWB)",
        governs="Excursion baselines, and the measured uranium a district's band is "
                "computed from.",
        stale_marks=("excursion_baselines",),
    ),
    "groundwater_levels": DatasetFile(
        key="groundwater_levels",
        relpath="cgwb_waterlevel_jharkhand.csv",
        kind="csv",
        id_column="id",
        label="Groundwater levels (CGWB)",
        governs="The groundwater flow field — hydraulic gradient and flow azimuth at "
                "every pin. Requires a flow-field rebuild to take effect.",
        stale_marks=("flow_field",),
    ),
    "ore_deposits": DatasetFile(
        key="ore_deposits",
        relpath="Jharkhand Ore/jharkhand_uranium_deposits.csv",
        kind="csv",
        id_column="name",
        label="Uranium deposit outlines",
        governs="Whether a pin resolves to deposit / belt / none — and therefore "
                "whether a uranium plume is possible there at all.",
    ),
    "ore_grades": DatasetFile(
        key="ore_grades",
        relpath="udepo_uranium_deposits.xlsx",
        kind="xlsx",
        id_column="Deposit ID",
        label="Uranium deposit grades (UDEPO)",
        governs="The grade factor scaling source concentration C0.",
        header_row=8,
    ),
    "naquim_vertical": DatasetFile(
        key="naquim_vertical",
        relpath="naquim_reference/naquim_vertical.csv",
        kind="csv",
        id_column="district",
        label="NAQUIM vertical structure",
        governs="Per-district aquifer layer bases and fracture depth ranges.",
    ),
}


def get(key: str) -> DatasetFile:
    try:
        return REGISTRY[key]
    except KeyError:
        raise DatasetError(f"unknown dataset '{key}'", status_code=404) from None


# ── reading ──────────────────────────────────────────────────────────

#: Byte-order mark. Written by `utf-8-sig`, and NOT present in most of these
#: files — see `_csv_encoding`.
_BOM = b"\xef\xbb\xbf"

#: Explicit, because `csv.writer` terminates lines with CRLF on every platform
#: regardless of the OS — which would rewrite every line of a file that shipped
#: with LF, turning an append into a whole-file diff.
LF = chr(10)


def _csv_encoding(p: Path) -> str:
    """Whatever encoding this file already uses, so a write cannot change it.

    `utf-8-sig` on WRITE emits a BOM unconditionally. Using it for every file
    added a BOM to `jharkhand_uranium_deposits.csv`, which shipped without one —
    so the file differed from HEAD forever after the first sync, even with every
    added row stripped back out. Only `waterQuality_jharkhand.csv` genuinely has
    a BOM; it must keep it, and the others must not gain one.
    """
    try:
        return "utf-8-sig" if p.read_bytes()[:3] == _BOM else "utf-8"
    except OSError:
        return "utf-8"


def _read(df_file: DatasetFile):
    """Load a dataset into a DataFrame with `record_source` guaranteed present."""
    import pandas as pd

    p = df_file.path
    if not p.exists():
        raise DatasetError(f"dataset file missing: {p}", status_code=404)

    if df_file.kind == "xlsx":
        df = pd.read_excel(p, header=df_file.header_row).dropna(how="all")
    else:
        # Read as utf-8-sig regardless: it strips a BOM if there is one — which
        # would otherwise become part of the first column's name ("﻿S. No.")
        # and break every lookup against `id_column` — and is a no-op if there is
        # not. Writing is the direction that must preserve the original.
        df = pd.read_csv(p, encoding="utf-8-sig")

    if SOURCE_COL not in df.columns:
        df[SOURCE_COL] = SOURCE_ORIGINAL
    else:
        df[SOURCE_COL] = df[SOURCE_COL].fillna(SOURCE_ORIGINAL)
    if REF_COL not in df.columns:
        df[REF_COL] = ""
    else:
        df[REF_COL] = df[REF_COL].fillna("")
    return df


# ── raw CSV access: values byte-for-byte as written ──────────────────
#
# WHY NOT PANDAS FOR WRITES. Round-tripping a CSV through pandas silently
# reformats every number: `22.7332550` came back as `22.733255`, `0.00` as
# `0.0`, and an int column containing one blank became floats. The values are
# numerically identical and the file is not — so after the very first sync,
# `jharkhand_uranium_deposits.csv` differed from the committed version on 8 rows
# nobody had touched, and stayed that way even after every added row was stripped
# back out. On a project whose entire claim is provenance, quietly rewriting the
# shipped evidence base is not an acceptable side effect of appending to it.
#
# `csv` hands back the exact strings and writes them back unchanged, so an
# untouched row survives a write untouched. pandas is still used for the
# READ-ONLY views (summary, paging, filtering), where reformatting is harmless.

def _read_raw(f: DatasetFile) -> tuple[list[str], list[list[str]], str]:
    """(header, rows, encoding) with every value exactly as it appears on disk."""
    import csv as _csv

    enc = _csv_encoding(f.path)
    with io.open(f.path, "r", encoding="utf-8-sig", newline="") as fh:
        rows = list(_csv.reader(fh))
    if not rows:
        raise DatasetError(f"{f.relpath} is empty", status_code=409)
    return rows[0], rows[1:], enc


def _write_raw(f: DatasetFile, header: list[str], rows: list[list[str]],
               enc: str) -> None:
    """Write header + rows back verbatim, preserving the original encoding.

    csv.writer defaults to CRLF on every platform, which would rewrite every
    line of a file that shipped with LF — hence the explicit LF terminator.
    """
    import csv as _csv

    with io.open(f.path, "w", encoding=enc, newline="") as fh:
        w = _csv.writer(fh, lineterminator=LF)
        w.writerow(header)
        w.writerows(rows)


def _col(header: list[str], name: str) -> int:
    """Index of a column, or -1."""
    for i, h in enumerate(header):
        if h.strip().lstrip("﻿") == name:
            return i
    return -1


def append_rows(f: DatasetFile, new: list[dict[str, Any]], ref: str) -> int:
    """Append rows to a CSV without touching a single existing byte.

    Adds `record_source` / `record_ref` to the header on first use and backfills
    the existing rows as `original` — that part does rewrite them, but only by
    adding two empty fields, never by reformatting a value.
    """
    # Before the first byte changes, not after: taking the snapshot later
    # captured a file that already had the provenance columns, so "restore what
    # shipped" restored the columns too.
    ensure_pristine(f)

    header, rows, enc = _read_raw(f)
    si, ri = _col(header, SOURCE_COL), _col(header, REF_COL)
    if si < 0:
        header.append(SOURCE_COL)
        for r in rows:
            r.append(SOURCE_ORIGINAL)
        si = len(header) - 1
    if ri < 0:
        header.append(REF_COL)
        for r in rows:
            r.append("")
        ri = len(header) - 1

    for item in new:
        row = [""] * len(header)
        for k, v in item.items():
            i = _col(header, str(k))
            if i >= 0:
                row[i] = "" if v is None else str(v)
        row[si] = SOURCE_ADDED
        row[ri] = ref
        rows.append(row)

    _write_raw(f, header, rows, enc)
    return len(new)


def _write_csv(df_file: DatasetFile, df) -> None:
    """Persist a DataFrame back over a CSV, preserving its original encoding."""
    df.to_csv(df_file.path, index=False, encoding=_csv_encoding(df_file.path))


# ── xlsx: mutate the sheet in place, never re-dump it ─────────────────
#
# The obvious implementation — read with pandas, write the frame back — was
# built first and silently corrupted the file: `grade_c0_factor("Jaduguda")`
# went from (0.6, 0.03) to (1.0, None) after a *no-op* round trip, with the row
# count unchanged. `pd.read_excel(header=8)` renames blank and duplicate header
# cells ("Unnamed: 5"), so writing back by matching the Excel header text against
# frame column names blanks every column whose name pandas altered.
#
# Nothing here reconstructs the sheet. It finds a row, changes or removes it, and
# leaves every other cell — and the 8-row preamble `ore_grades.py` relies on —
# exactly as it was.

def _xlsx_open(f: DatasetFile):
    import openpyxl

    wb = openpyxl.load_workbook(f.path)
    ws = wb[wb.sheetnames[0]]
    hrow = f.header_row + 1                      # openpyxl is 1-indexed
    header = [c.value for c in ws[hrow]]
    return wb, ws, hrow, header


def _xlsx_ensure_source(f: DatasetFile) -> None:
    """Add `record_source` to the sheet and backfill it, once."""
    ensure_pristine(f)
    wb, ws, hrow, header = _xlsx_open(f)
    present = [str(h) for h in header if h is not None]
    if SOURCE_COL in present and REF_COL in present:
        return
    # `max_column + 1`, NOT `len(non-empty headers) + 1`. This sheet's used range
    # starts at column B, so counting names put the new header straight on top of
    # `Grade Range` — which silently turned grade_c0_factor("Jaduguda") from
    # (0.6, 0.03) into (1.0, None) with no error and no row-count change.
    filled = 0
    for name, default in ((SOURCE_COL, SOURCE_ORIGINAL), (REF_COL, "")):
        # Re-read: adding a cell above moves `max_column`, and the second
        # column must land after the first, not on top of it.
        header = [c.value for c in ws[hrow]]
        if name in [str(h) for h in header if h is not None]:
            continue
        col = ws.max_column + 1
        ws.cell(row=hrow, column=col, value=name)
        for r in range(hrow + 1, ws.max_row + 1):
            if any(ws.cell(row=r, column=c).value is not None
                   for c in range(1, col)):
                ws.cell(row=r, column=col, value=default)
                filled += 1
    wb.save(f.path)
    logger.info(f"{f.relpath}: added '{SOURCE_COL}', backfilled {filled} row(s) "
                f"as '{SOURCE_ORIGINAL}'")


def _xlsx_col_index(header: list, name: str) -> int:
    for i, h in enumerate(header, start=1):
        if h is not None and str(h).strip() == name:
            return i
    raise DatasetError(f"column {name!r} not found in sheet header")


def _xlsx_locate(ws, hrow: int, header: list, f: DatasetFile, row_id: str) -> int:
    """Excel row number for one natural-key value."""
    idc = _xlsx_col_index(header, f.id_column)
    hits = [r for r in range(hrow + 1, ws.max_row + 1)
            if str(ws.cell(row=r, column=idc).value or "").strip() == str(row_id).strip()]
    if not hits:
        raise DatasetError(
            f"no row with {f.id_column}={row_id!r} in {f.relpath}", status_code=404)
    if len(hits) > 1:
        raise DatasetError(
            f"{f.id_column}={row_id!r} matches {len(hits)} rows in {f.relpath}; "
            f"refusing to act on an ambiguous key")
    return hits[0]


def _xlsx_row_source(ws, excel_row: int, header: list) -> str:
    try:
        sc = _xlsx_col_index(header, SOURCE_COL)
    except DatasetError:
        return SOURCE_ORIGINAL
    return str(ws.cell(row=excel_row, column=sc).value or SOURCE_ORIGINAL)


def summary() -> list[dict[str, Any]]:
    """One row per writable dataset: counts, split, and what it governs."""
    out = []
    for key, f in REGISTRY.items():
        entry: dict[str, Any] = {
            "key": key, "label": f.label, "path": f.relpath,
            "kind": f.kind, "id_column": f.id_column, "governs": f.governs,
        }
        try:
            df = _read(f)
            counts = df[SOURCE_COL].value_counts().to_dict()
            entry.update({
                "rows": int(len(df)),
                "original": int(counts.get(SOURCE_ORIGINAL, 0)),
                "added": int(counts.get(SOURCE_ADDED, 0)),
                "modified_at": datetime.fromtimestamp(
                    f.path.stat().st_mtime, tz=timezone.utc).isoformat(),
                "available": True,
            })
        except Exception as exc:  # noqa: BLE001
            entry.update({"available": False, "error": str(exc)})
        out.append(entry)
    return out


def rows(key: str, *, source: Optional[str] = None, q: Optional[str] = None,
         offset: int = 0, limit: int = 100) -> dict[str, Any]:
    """Paged rows, newest-added first so a just-synced row is visible."""
    f = get(key)
    df = _read(f)
    df = df.reset_index(drop=True)

    if source in (SOURCE_ORIGINAL, SOURCE_ADDED):
        df = df[df[SOURCE_COL] == source]
    if q:
        needle = str(q).strip().lower()
        mask = df.apply(
            lambda r: needle in " ".join(str(v).lower() for v in r.values), axis=1)
        df = df[mask]

    # `added` rows first: they are the ones an admin came here to act on.
    df = df.assign(_added=(df[SOURCE_COL] == SOURCE_ADDED).astype(int))
    df = df.sort_values("_added", ascending=False, kind="stable").drop(columns="_added")

    total = int(len(df))
    page = df.iloc[offset:offset + limit]
    import pandas as pd
    recs = [
        {k: (None if pd.isna(v) else v) for k, v in rec.items()}
        for rec in page.to_dict(orient="records")
    ]
    return {
        "key": key, "label": f.label, "id_column": f.id_column,
        "columns": [c for c in df.columns],
        "total": total, "offset": offset, "limit": limit,
        "editable_note": (
            f"Rows marked '{SOURCE_ORIGINAL}' are the shipped evidence base and "
            f"cannot be edited or deleted. Only '{SOURCE_ADDED}' rows — written by "
            f"this system from approved field observations — may be changed."),
        "rows": recs,
    }


# ── writing ──────────────────────────────────────────────────────────

#: Backups live here, NOT beside the original.
#:
#: Writing `foo.csv.<ref>.bak` next to `foo.csv` put untracked litter straight
#: into a tracked directory: four of them accumulated in `Datasets/` during
#: testing and showed up in every `git status`, which is exactly the noise that
#: makes a real change hard to spot. Dot-prefixed and gitignored.
BACKUP_DIR = DATASETS / ".backups"

#: A copy of each file as it was BEFORE this system first wrote to it.
#:
#: Taken once, never overwritten, and used by the reset to guarantee "back to
#: what shipped" means byte-for-byte. It exists because faithfulness cannot be
#: assumed from the write path: the CSV writers were made byte-preserving, but
#: `openpyxl` re-serialises an entire workbook on save — new zip member order,
#: rewritten metadata, dropped formatting — so `udepo_uranium_deposits.xlsx`
#: could never round-trip identically no matter how carefully rows were edited.
#: Restoring a snapshot sidesteps the whole class of problem.
PRISTINE_DIR = BACKUP_DIR / "pristine"


def _pristine_path(f: DatasetFile) -> Path:
    return PRISTINE_DIR / f.relpath


def is_pristine_snapshot(f: DatasetFile) -> bool:
    """Is the stored snapshot genuinely the shipped file?

    A snapshot is only meaningful if it was taken BEFORE this system ever wrote.
    On a database that had already synced before snapshotting existed, the first
    `ensure_pristine` captures a file that already carries added rows and the
    provenance columns — and restoring THAT undoes a reset instead of completing
    it. (Observed: `strip_added` removed the rows, then the tail restored the
    snapshot and put them straight back.)

    So a snapshot is trusted only if it looks like something that shipped: no
    provenance columns, and therefore no added rows.
    """
    p = _pristine_path(f)
    if not p.exists():
        return False
    if f.kind == "xlsx":
        # Cheap structural check; the workbook is re-serialised on any write, so
        # the presence of the column is the reliable signal.
        try:
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True)
            ws = wb[wb.sheetnames[0]]
            header = [c.value for c in next(ws.iter_rows(
                min_row=f.header_row + 1, max_row=f.header_row + 1))]
            return SOURCE_COL not in [str(h) for h in header if h is not None]
        except Exception:  # noqa: BLE001
            return False
    try:
        with io.open(p, "r", encoding="utf-8-sig", newline="") as fh:
            first = fh.readline()
        return SOURCE_COL not in first
    except OSError:
        return False


def ensure_pristine(f: DatasetFile) -> None:
    """Snapshot the file the first time it is about to be written. Idempotent."""
    dest = _pristine_path(f)
    if dest.exists() or not f.path.exists():
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(f.path, dest)
    logger.info(f"{f.relpath}: pristine snapshot taken before first write")


def backup_file(f: DatasetFile, ref: str) -> Path:
    """Back up a registry file, taking its pristine snapshot on the first write."""
    ensure_pristine(f)
    return backup(f.path, ref)


def backup(path: Path, ref: str) -> Path:
    """Copy the file aside before rewriting.

    These are tracked data files. Any admin action that changes one should be
    trivially undoable without reaching for git, because the person undoing it
    may be doing so at speed.
    """
    dest_dir = BACKUP_DIR / path.parent.relative_to(DATASETS)
    dest_dir.mkdir(parents=True, exist_ok=True)
    bak = dest_dir / f"{path.name}.{ref}.bak"
    shutil.copy2(path, bak)
    return bak


def _locate(df, f: DatasetFile, row_id: str):
    """Find exactly one row by the file's natural key."""
    ids = df[f.id_column].astype(str).str.strip()
    match = df[ids == str(row_id).strip()]
    if match.empty:
        raise DatasetError(
            f"no row with {f.id_column}={row_id!r} in {f.relpath}", status_code=404)
    if len(match) > 1:
        raise DatasetError(
            f"{f.id_column}={row_id!r} matches {len(match)} rows in {f.relpath}; "
            f"refusing to act on an ambiguous key")
    return match.index[0], match.iloc[0]


def _raise_original(f: DatasetFile, row_id: str, verb: str) -> None:
    raise DatasetError(
        f"cannot {verb} {f.id_column}={row_id!r}: it is a '{SOURCE_ORIGINAL}' row. "
        f"The shipped evidence base (CGWB, UDEPO, GSI, NAQUIM) is immutable — only "
        f"rows this system added from approved field observations can be changed. To "
        f"correct source data, replace the file through /ingest, which records a "
        f"versioned, checksummed dataset.")


def _guard_original(row, f: DatasetFile, row_id: str, verb: str) -> None:
    if str(row.get(SOURCE_COL, SOURCE_ORIGINAL)) != SOURCE_ADDED:
        _raise_original(f, row_id, verb)


def _validate_patch(f: DatasetFile, columns, patch: dict[str, Any]) -> None:
    unknown = [k for k in patch if k not in columns]
    if unknown:
        raise DatasetError(f"unknown column(s) for {f.relpath}: {unknown}")
    if not patch:
        raise DatasetError("empty patch — nothing to change")
    if SOURCE_COL in patch:
        raise DatasetError(
            f"{SOURCE_COL} is provenance, not data — it cannot be edited. A row "
            f"cannot be relabelled as shipped evidence.")
    if f.id_column in patch:
        raise DatasetError(
            f"{f.id_column} is this file's row identity and cannot be edited; "
            f"delete the row and re-sync instead.")


def update_row(key: str, row_id: str, patch: dict[str, Any]) -> dict[str, Any]:
    """Edit one `added` row. Returns before/after for the audit entry."""
    f = get(key)
    ref = uuid.uuid4().hex[:12]

    if f.kind == "xlsx":
        _xlsx_ensure_source(f)
        wb, ws, hrow, header = _xlsx_open(f)
        _validate_patch(f, [str(h) for h in header if h is not None], patch)
        r = _xlsx_locate(ws, hrow, header, f, row_id)
        if _xlsx_row_source(ws, r, header) != SOURCE_ADDED:
            _raise_original(f, row_id, "edit")
        bak = backup_file(f, ref)
        before, after = {}, {}
        for col, val in patch.items():
            c = _xlsx_col_index(header, col)
            before[col] = ws.cell(row=r, column=c).value
            ws.cell(row=r, column=c, value=val)
            after[col] = val
        wb.save(f.path)
    else:
        df = _read(f)
        _validate_patch(f, df.columns, patch)
        idx, row = _locate(df, f, row_id)
        _guard_original(row, f, row_id, "edit")
        before = {k: (None if _isna(row[k]) else row[k]) for k in patch}
        for k, v in patch.items():
            df.at[idx, k] = v
        after = {k: (None if _isna(df.at[idx, k]) else df.at[idx, k]) for k in patch}
        bak = backup_file(f, ref)
        _write_csv(f, df)

    invalidate_caches()
    logger.info(f"{f.relpath}: edited {f.id_column}={row_id} ({len(patch)} field(s))")
    return {"ref": ref, "backup": str(bak.relative_to(REPO_ROOT)),
            "before": _jsonable(before), "after": _jsonable(after),
            "stale_marks": list(f.stale_marks)}


def delete_row(key: str, row_id: str) -> dict[str, Any]:
    """Remove one `added` row. Returns the removed row for the audit entry."""
    f = get(key)
    ref = uuid.uuid4().hex[:12]

    if f.kind == "xlsx":
        _xlsx_ensure_source(f)
        wb, ws, hrow, header = _xlsx_open(f)
        r = _xlsx_locate(ws, hrow, header, f, row_id)
        if _xlsx_row_source(ws, r, header) != SOURCE_ADDED:
            _raise_original(f, row_id, "delete")
        removed = {str(h): ws.cell(row=r, column=i).value
                   for i, h in enumerate(header, start=1) if h is not None}
        bak = backup_file(f, ref)
        ws.delete_rows(r)
        wb.save(f.path)
    else:
        df = _read(f)
        idx, row = _locate(df, f, row_id)
        _guard_original(row, f, row_id, "delete")
        removed = {k: (None if _isna(v) else v) for k, v in row.items()}
        bak = backup_file(f, ref)
        _write_csv(f, df.drop(index=idx))

    invalidate_caches()
    logger.info(f"{f.relpath}: deleted {f.id_column}={row_id}")
    return {"ref": ref, "backup": str(bak.relative_to(REPO_ROOT)),
            "removed": _jsonable(removed),
            "record_ref": (str(removed.get(REF_COL) or "").strip() or None),
            "stale_marks": list(f.stale_marks)}


def drop_provenance_columns(f: DatasetFile) -> bool:
    """Remove `record_source` / `record_ref` if present. True if anything changed.

    Split out because the columns must come off in BOTH strip paths. They used to
    be dropped only when there was at least one `added` row to remove, so a file
    that had gained the columns and then had its added rows deleted one at a time
    kept them for ever — still differing from what shipped, still showing in
    `git status`, and still failing `is_pristine_snapshot`.
    """
    if f.kind == "xlsx":
        wb, ws, hrow, header = _xlsx_open(f)
        changed = False
        for name in (REF_COL, SOURCE_COL):
            header = [c.value for c in ws[hrow]]
            for i, h in enumerate(header, start=1):
                if h is not None and str(h).strip() == name:
                    ws.delete_cols(i)
                    changed = True
                    break
        if changed:
            wb.save(f.path)
        return changed

    header, rows, enc = _read_raw(f)
    changed = False
    for name in (REF_COL, SOURCE_COL):
        i = _col(header, name)
        if i >= 0:
            header.pop(i)
            for r in rows:
                if i < len(r):
                    r.pop(i)
            changed = True
    if changed:
        _write_raw(f, header, rows, enc)
    return changed


def strip_added(key: str, *, dry_run: bool = False) -> dict[str, Any]:
    """Remove every `added` row — the per-file primitive behind the reset."""
    f = get(key)
    df = _read(f)
    added = df[df[SOURCE_COL] == SOURCE_ADDED]
    result: dict[str, Any] = {
        "key": key, "path": f.relpath, "would_remove": int(len(added)),
        "ids": [str(v) for v in added[f.id_column].tolist()],
        "record_refs": sorted({str(v).strip() for v in added[REF_COL].tolist()
                               if str(v).strip()}),
    }
    if dry_run:
        return result

    if added.empty:
        # No rows to strip — but "reset to what shipped" still means the file
        # ends up as it shipped. A file whose columns were added and whose rows
        # were later deleted one by one has nothing left to strip and is still
        # not pristine.
        pristine = _pristine_path(f)
        if is_pristine_snapshot(f) and f.path.read_bytes() != pristine.read_bytes():
            backup(f.path, uuid.uuid4().hex[:12])
            shutil.copy2(pristine, f.path)
            invalidate_caches()
            result["restored_from_pristine"] = True
            result["removed"] = 0
            logger.info(f"{f.relpath}: no added rows; restored from pristine")
        elif drop_provenance_columns(f):
            invalidate_caches()
            result["provenance_columns_dropped"] = True
            result["removed"] = 0
            logger.info(f"{f.relpath}: no added rows; dropped the provenance columns")
        return result

    ref = uuid.uuid4().hex[:12]
    result["backup"] = str(backup_file(f, ref).relative_to(REPO_ROOT))
    if f.kind == "xlsx":
        wb, ws, hrow, header = _xlsx_open(f)
        # Bottom-up: deleting a row shifts every row beneath it.
        for r in sorted(
            (rr for rr in range(hrow + 1, ws.max_row + 1)
             if _xlsx_row_source(ws, rr, header) == SOURCE_ADDED), reverse=True):
            ws.delete_rows(r)
        wb.save(f.path)
        drop_provenance_columns(f)
    else:
        # Removing the rows is not enough to restore the file.
        #
        # `record_source` and `record_ref` are added on first write and stay
        # forever, so a file that had every added row stripped back out STILL
        # differed from the shipped version — two extra header columns, and a
        # permanent `git status` entry that made it look as though the reset had
        # not worked. It had; the file's *shape* had changed.
        #
        # With no added rows left there is nothing for the columns to describe,
        # so they come off and the file goes back to exactly what shipped —
        # byte for byte, which is why this filters raw rows rather than
        # rewriting a DataFrame.
        header, raw, enc = _read_raw(f)
        si = _col(header, SOURCE_COL)
        keep = [r for r in raw
                if not (si >= 0 and si < len(r) and r[si] == SOURCE_ADDED)]
        for name in (REF_COL, SOURCE_COL):
            i = _col(header, name)
            if i >= 0:
                header.pop(i)
                for r in keep:
                    if i < len(r):
                        r.pop(i)
        _write_raw(f, header, keep, enc)
    # If a pristine snapshot exists and nothing added survives, put the snapshot
    # back. Row-filtering already produced the right CONTENT; this guarantees the
    # right BYTES, which is what "restored to what shipped" has to mean for a
    # project whose claim is provenance.
    if is_pristine_snapshot(f):
        shutil.copy2(_pristine_path(f), f.path)
        result["restored_from_pristine"] = True

    invalidate_caches()
    result["removed"] = int(len(added))
    result["provenance_columns_dropped"] = True
    logger.info(f"{f.relpath}: stripped {len(added)} added row(s)"
                + ("; restored byte-for-byte from the pristine snapshot"
                   if result.get("restored_from_pristine") else ""))
    return result


def list_backups(key: str) -> list[dict[str, Any]]:
    f = get(key)
    out = []
    bdir = BACKUP_DIR / f.path.parent.relative_to(DATASETS)
    if not bdir.exists():
        return out
    for b in sorted(bdir.glob(f"{f.path.name}.*.bak"),
                    key=lambda p: p.stat().st_mtime, reverse=True):
        out.append({
            "name": b.name,
            "ref": b.name.rsplit(".", 2)[-2],
            "size_bytes": b.stat().st_size,
            "created_at": datetime.fromtimestamp(
                b.stat().st_mtime, tz=timezone.utc).isoformat(),
        })
    return out


def restore(key: str, backup_name: str) -> dict[str, Any]:
    """Put a backup back. The current file is itself backed up first."""
    f = get(key)
    bdir = (BACKUP_DIR / f.path.parent.relative_to(DATASETS)).resolve()
    src = (bdir / backup_name).resolve()
    # Contain the path: `backup_name` arrives from the client.
    if (src.parent != bdir or not src.name.endswith(".bak")
            or "/" in backup_name or "\\" in backup_name):
        raise DatasetError(f"not a backup of {f.relpath}: {backup_name!r}")
    if not src.exists():
        raise DatasetError(f"no such backup: {backup_name}", status_code=404)

    ref = uuid.uuid4().hex[:12]
    pre = backup(f.path, ref)
    shutil.copy2(src, f.path)
    invalidate_caches()
    logger.info(f"{f.relpath}: restored from {backup_name}")
    return {"restored_from": backup_name, "ref": ref,
            "backup_of_previous": str(pre.relative_to(REPO_ROOT)),
            "stale_marks": list(f.stale_marks)}


# ── helpers ──────────────────────────────────────────────────────────

def _isna(v) -> bool:
    import pandas as pd
    try:
        return bool(pd.isna(v))
    except (TypeError, ValueError):
        return False


def _jsonable(d: dict[str, Any]) -> dict[str, Any]:
    """numpy scalars are not JSON-serialisable and land in audit detail."""
    out = {}
    for k, v in d.items():
        if hasattr(v, "item"):
            try:
                v = v.item()
            except (AttributeError, ValueError):
                v = str(v)
        out[str(k)] = v
    return out


def invalidate_caches() -> None:
    """Drop the pipeline's memoised dataset reads.

    `ore_loader`, `ore_grades` and `jharkhand_loader` memoise their parsed files
    with `functools.lru_cache`, so without this a running process keeps serving
    pre-edit data until it restarts — and the edit would look like it had done
    nothing.
    """
    import sys
    if str(REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(REPO_ROOT))
    cleared = []
    for mod_name in ("ml_pipeline.data_prep.ore_loader",
                     "ml_pipeline.data_prep.ore_grades",
                     "ml_pipeline.data_prep.jharkhand_loader"):
        try:
            mod = __import__(mod_name, fromlist=["*"])
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"could not reach {mod_name} to clear its cache: {exc}")
            continue
        for attr in dir(mod):
            fn = getattr(mod, attr)
            if hasattr(fn, "cache_clear"):
                fn.cache_clear()
                cleared.append(f"{mod_name}.{attr}")
    logger.info(f"cleared {len(cleared)} ml_pipeline dataset cache(s)")
