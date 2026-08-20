"""Carry approved field observations into `Datasets/`, deliberately and by hand.

WHY THIS IS MANUAL. Three architectures were considered (PRODUCT_DESIGN.md
§3.6): auto-feeding the database into the engine, a continuous DB→dataset sync,
or an admin-triggered export. The first breaks reproducibility — the same pin
and sliders would return different answers over time with nothing recorded, and
a regulator cannot defend a number that moves. The second is the right end state
if this ever ingests sensor streams, and is heavy machinery for the handful of
ore sightings actually expected. So: **admin-triggered, audited, reversible.**

ALL THREE TYPES SYNC (R11). Ore was automated first because it is the one input
that otherwise does nothing at all: `ore_zone_at()` reads the deposit CSV, and a
pin in `zone == "none"` suppresses the uranium plume entirely (frozen rule #3,
"the tool cannot invent contamination"), so an approved *"uranium ore found here"*
left the simulation still reporting no plume. Chemistry and groundwater levels
were left manual on the argument that they only "move a feature value the model
was already trained across" — true of the *surrogate*, but not of the serve-time
inputs derived from those files, and the backlog simply accumulated.

WHAT A SYNC TOUCHES

    Datasets/Jharkhand Ore/jharkhand_uranium_deposits.csv
        Drives `ore_zone_at()` — whether a pin is deposit / belt / none, and
        therefore whether a uranium plume is possible at all.
    Datasets/udepo_uranium_deposits.xlsx  (header row 8)
        Drives `grade_c0_factor()` — scales the source concentration C0.
    Datasets/waterQuality_jharkhand.csv
        Drives the ambient excursion baselines. **Derived** — needs a baseline
        recompute before the engine sees it.
    Datasets/cgwb_waterlevel_jharkhand.csv
        Drives the groundwater flow field. **Derived** — needs a flow-field
        rebuild (which needs the GLO-30 DEM) before the engine sees it.

Every file gains a `record_source` column — `original` for rows that shipped with
the project, `added` for anything an admin approved — backfilled on first touch.
`app/services/datasets.py` owns that column and the row-level edit/delete path
built on it, including the rule that `original` rows are immutable. It is not
called `source` because two of these files already have a `source` column meaning
something else entirely (collecting agency; citation).

THIS DOES NOT RETRAIN ANYTHING. Adding a deposit changes a *resolved input*, not
the model: C0 and the ore zone are read at serve time, and the surrogate was
trained across the full Texas C0 envelope. Retraining is only required when the
generator's assumptions change (§4.6 rule 9). What a sync CAN invalidate is a
derived artifact, so each result carries `stale_marks` naming what to rebuild —
see `/model-ops/status`.
"""
from __future__ import annotations

import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from loguru import logger
from sqlalchemy import bindparam as sa_bindparam, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import AppException
from app.services import audit, jobs

REPO_ROOT = Path(__file__).resolve().parents[3]
ORE_CSV = REPO_ROOT / "Datasets" / "Jharkhand Ore" / "jharkhand_uranium_deposits.csv"
UDEPO_XLSX = REPO_ROOT / "Datasets" / "udepo_uranium_deposits.xlsx"
UDEPO_HEADER_ROW = 8          # matches ml_pipeline/data_prep/ore_grades.py

# The provenance column lives in `datasets.py`, which owns the immutability rule
# that depends on it. Re-exported here under the old names so the ore writer
# below reads unchanged; there is exactly one column, in one place.
from app.services import datasets as dsx  # noqa: E402

ORIGIN_COL = dsx.SOURCE_COL
ORIGIN_ORIGINAL = dsx.SOURCE_ORIGINAL
ORIGIN_ADDED = dsx.SOURCE_ADDED

#: All three observation types now have an automated path. Chemistry and
#: groundwater levels were manual until R11 on the reasoning that they "move a
#: feature value the model was already trained across" — true of the surrogate,
#: but not of the serve-time inputs derived from those files. A water-quality row
#: shifts the excursion baselines a UCL is computed against; a groundwater-level
#: row shifts the flow field that sets gradient and azimuth at every pin. Leaving
#: them to be applied by hand from the audit log meant the portal and the engine
#: disagreed indefinitely, which is the split-brain the amber badge was reporting
#: rather than fixing.
SYNCABLE_TYPES = ("ore_presence", "water_sample", "groundwater_level")


class DatasetSyncError(AppException):
    def __init__(self, message: str):
        super().__init__(message, status_code=409)


def _radius_polygon_wkt(lon: float, lat: float, radius_m: float = 400.0,
                        n: int = 24) -> str:
    """A small circular outline around the sighting.

    A field observation is a point; `ore_zone_at()` needs an outline to test
    containment against. 400 m is the order of the existing deposit outlines and
    is recorded in the row's notes so nobody mistakes it for a surveyed boundary.
    """
    import math
    dlat = radius_m / 111_320.0
    dlon = radius_m / (111_320.0 * max(math.cos(math.radians(lat)), 1e-6))
    pts = []
    for i in range(n + 1):
        th = 2.0 * math.pi * i / n
        pts.append(f"{lon + dlon * math.cos(th):.6f} {lat + dlat * math.sin(th):.6f}")
    return f"POLYGON(({', '.join(pts)}))"


async def pending_summary(db: AsyncSession) -> dict[str, Any]:
    """Counts for the three UI states, plus the amber backlog by type."""
    rows = (await db.execute(text("""
        SELECT observation_type,
               count(*) FILTER (WHERE status = 'pending')                       AS pending_review,
               count(*) FILTER (WHERE status = 'approved'
                                  AND synced_to_dataset_at IS NULL)             AS approved_unsynced,
               count(*) FILTER (WHERE status = 'approved'
                                  AND synced_to_dataset_at IS NOT NULL)         AS in_model
        FROM field_observations
        GROUP BY observation_type
    """))).mappings().all()

    # ORPHANS: approved, unsynced, and the row approval created no longer
    # exists. The sync joins to that row, so it matches nothing and reports
    # "Nothing to sync" — for ever — while this summary keeps counting the
    # observation as pending. That is a permanently stuck split-brain, and it is
    # exactly the state the sync status exists to make impossible.
    orphan_rows = (await db.execute(text("""
        SELECT f.observation_type, count(*) AS n
        FROM field_observations f
        WHERE f.status = 'approved' AND f.synced_to_dataset_at IS NULL
          AND CASE f.observation_type
                WHEN 'water_sample' THEN
                  NOT EXISTS (SELECT 1 FROM water_samples w
                              WHERE w.id = COALESCE(f.applied_id, f.target_id))
                WHEN 'groundwater_level' THEN
                  NOT EXISTS (SELECT 1 FROM groundwater_level_readings r
                              WHERE r.id = COALESCE(f.applied_id, f.target_id))
                ELSE false END
        GROUP BY f.observation_type
    """))).mappings().all()
    orphans = {r["observation_type"]: int(r["n"]) for r in orphan_rows}
    orphan_total = sum(orphans.values())

    by_type = {r["observation_type"]: dict(r) for r in rows}
    total_unsynced = sum(r["approved_unsynced"] for r in rows)
    total_pending = sum(r["pending_review"] for r in rows)
    total_in_model = sum(r["in_model"] for r in rows)

    return {
        "pending_review": total_pending,
        "approved_pending_sync": total_unsynced,
        "approved_in_model": total_in_model,
        "by_type": by_type,
        "orphaned": orphan_total,
        "orphaned_by_type": orphans,
        "orphan_note": (
            f"{orphan_total} approved observation(s) can no longer be synced: the "
            f"record they created has since been deleted, so there is nothing left "
            f"to carry into the datasets. They will never clear on their own — "
            f"resolve them with POST /dataset-sync/reconcile, then re-submit if "
            f"the observation is still wanted."
            if orphan_total else None),
        "syncable_types": list(SYNCABLE_TYPES),
        # The sentence the UI shows verbatim.
        "message": (
            (f"{total_unsynced - orphan_total} approved observation(s) are not yet "
             f"in the model" + (f"; {orphan_total} more cannot be synced at all."
                                if orphan_total else "."))
            if total_unsynced else
            "All approved observations are reflected in the model."),
        "note": ("Approved observations are authoritative in the portal "
                 "immediately, but the physics engine and surrogate read only "
                 "Datasets/. A sync is a deliberate admin action; it changes "
                 "resolved inputs, not the trained model."),
    }


async def _unsynced_ore(db: AsyncSession) -> list[dict[str, Any]]:
    rows = (await db.execute(text("""
        SELECT f.id::text            AS obs_id,
               o.id::text            AS ore_id,
               o.name, o.ore_zone, o.uranium_grade_pct, o.depth_m,
               o.radius_m, o.notes, o.observed_at,
               ST_X(o.location::geometry) AS lon,
               ST_Y(o.location::geometry) AS lat
        FROM field_observations f
        JOIN ore_observations o ON o.origin_observation_id = f.id
        WHERE f.observation_type = 'ore_presence'
          AND f.status = 'approved'
          AND f.synced_to_dataset_at IS NULL
        ORDER BY f.reviewed_at
    """))).mappings().all()
    return [dict(r) for r in rows]


def _ensure_origin_column_csv(df, path: Path) -> None:
    if ORIGIN_COL not in df.columns:
        df[ORIGIN_COL] = ORIGIN_ORIGINAL
        logger.info(f"{path.name}: added '{ORIGIN_COL}' column, "
                    f"backfilled {len(df)} row(s) as '{ORIGIN_ORIGINAL}'")
    else:
        df[ORIGIN_COL] = df[ORIGIN_COL].fillna(ORIGIN_ORIGINAL)


def _backup(path: Path, ref: str) -> Path:
    """Copy beside the original before rewriting. These files are tracked data;
    an admin action that appends to them should be trivially undoable."""
    bak = path.with_suffix(path.suffix + f".{ref}.bak")
    shutil.copy2(path, bak)
    return bak


async def sync_ore(db: AsyncSession, *, actor, dry_run: bool = False,
                   ip: Optional[str] = None) -> dict[str, Any]:
    """Append approved ore observations to both dataset files."""
    import pandas as pd

    items = await _unsynced_ore(db)
    ref = uuid.uuid4().hex[:12]
    result: dict[str, Any] = {
        "sync_ref": ref, "dry_run": dry_run, "count": len(items),
        "deposits": [i["name"] for i in items],
        "files": [], "backups": [],
        "retrain_required": False,
        "note": ("Adding a deposit changes a RESOLVED INPUT (ore zone and C0), "
                 "not the trained model. No retrain or re-bake is required; the "
                 "surrogate was trained across the full source envelope. See "
                 "PRODUCT_DESIGN.md section 4.6 rule 9 for what does require one."),
    }
    if not items:
        result["message"] = "Nothing to sync."
        return result

    job_id = None if dry_run else jobs.start(
        "sync_ore", label="Syncing approved ore observations", actor=getattr(actor, "email", None))

    for p in (ORE_CSV, UDEPO_XLSX):
        if not p.exists():
            raise DatasetSyncError(f"dataset file missing: {p}")

    # ── the deposit outline CSV — drives ore_zone_at() ───────────────
    csv_df = pd.read_csv(ORE_CSV)
    _ensure_origin_column_csv(csv_df, ORE_CSV)
    existing = {str(n).strip().lower() for n in csv_df["name"].fillna("")}

    new_csv_rows, skipped = [], []
    for it in items:
        if str(it["name"]).strip().lower() in existing:
            # Name collision: ore_zone_at and grade lookup both key on name, so
            # a duplicate would make the grade ambiguous. Refuse the row rather
            # than shadow an existing deposit.
            skipped.append(it["name"])
            continue
        new_csv_rows.append({
            "name": it["name"],
            "district": "", "state": "Jharkhand",
            "center_lat": round(float(it["lat"]), 6),
            "center_lon": round(float(it["lon"]), 6),
            "status": "Field-observed",
            "geometry_wkt": _radius_polygon_wkt(
                float(it["lon"]), float(it["lat"]),
                radius_m=float(it["radius_m"]) if it.get("radius_m") else 400.0),
            "notes": (f"Field observation approved {it['observed_at']}. "
                      f"Outline is a "
                      f"{float(it['radius_m']) if it.get('radius_m') else 400.0:g} m "
                      f"radius {'as reported by the submitter' if it.get('radius_m') else 'DEFAULT — no extent was reported'}"
                      f", NOT a surveyed boundary. "
                      f"{(it['notes'] or '').strip()}").strip(),
            ORIGIN_COL: ORIGIN_ADDED,
            dsx.REF_COL: ref,
        })
    result["skipped_duplicate_name"] = skipped

    # ── the grade workbook — drives grade_c0_factor() ────────────────
    xl_df = pd.read_excel(UDEPO_XLSX, header=UDEPO_HEADER_ROW).dropna(how="all")
    if ORIGIN_COL not in xl_df.columns:
        xl_df[ORIGIN_COL] = ORIGIN_ORIGINAL
    else:
        xl_df[ORIGIN_COL] = xl_df[ORIGIN_COL].fillna(ORIGIN_ORIGINAL)

    new_xl_rows = []
    for it in items:
        if it["name"] in skipped:
            continue
        grade = it["uranium_grade_pct"]
        new_xl_rows.append({
            "Country": "India",
            "Deposit ID": f"FIELD-{it['ore_id'][:8]}",
            "Deposit Name": it["name"],
            "Main Commodity": "Uranium",
            "Deposit Type": "Field observation",
            "Deposit Subtype": it["ore_zone"],
            "Resource Range": "",
            # `_parse_grade` reads a range string; a single value is valid input.
            "Grade Range": (f"{float(grade):g}" if grade is not None else ""),
            ORIGIN_COL: ORIGIN_ADDED,
            dsx.REF_COL: ref,
        })

    if dry_run:
        result["message"] = (f"Would append {len(new_csv_rows)} deposit row(s) "
                             f"and {len(new_xl_rows)} grade row(s).")
        return result

    if new_csv_rows:
        result["backups"].append(str(dsx.backup_file(dsx.get("ore_deposits"), ref).relative_to(REPO_ROOT)))
        # Same reason as the chemistry writer: a pandas rewrite silently turned
        # 22.7332550 into 22.733255 on rows nobody had touched.
        dsx.append_rows(dsx.get("ore_deposits"), new_csv_rows, ref)
        result["files"].append(str(ORE_CSV.relative_to(REPO_ROOT)))

    if new_xl_rows:
        result["backups"].append(
            str(dsx.backup_file(dsx.get("ore_grades"), ref).relative_to(REPO_ROOT)))
        out = pd.concat([xl_df, pd.DataFrame(new_xl_rows)], ignore_index=True)
        # Preserve the 8-row preamble so ore_grades.py's `header=8` still lands
        # on the real header row.
        import openpyxl
        wb = openpyxl.load_workbook(UDEPO_XLSX)
        ws = wb[wb.sheetnames[0]]
        header = [c.value for c in ws[UDEPO_HEADER_ROW + 1]]
        if ORIGIN_COL not in [str(h) for h in header]:
            ws.cell(row=UDEPO_HEADER_ROW + 1, column=len(header) + 1,
                    value=ORIGIN_COL)
            for r in range(UDEPO_HEADER_ROW + 2, ws.max_row + 1):
                ws.cell(row=r, column=len(header) + 1, value=ORIGIN_ORIGINAL)
            header = header + [ORIGIN_COL]
        for row in new_xl_rows:
            ws.append([row.get(str(h), "") for h in header])
        wb.save(UDEPO_XLSX)
        result["files"].append(str(UDEPO_XLSX.relative_to(REPO_ROOT)))

    synced_ids = [it["obs_id"] for it in items if it["name"] not in skipped]
    if synced_ids:
        # `expanding=True` rather than a Postgres array literal: asyncpg binds
        # by Python type and rejects the '{a,b}' string form outright.
        stmt = text("""
            UPDATE field_observations
            SET synced_to_dataset_at = now(), dataset_sync_ref = :ref
            WHERE id IN :ids
        """).bindparams(sa_bindparam("ids", expanding=True))
        await db.execute(stmt, {"ref": ref,
                                "ids": [uuid.UUID(i) for i in synced_ids]})
        await db.commit()

    invalidate_ml_caches()

    await audit.record(
        action="dataset.sync_ore", entity_type="datasets", entity_id=ref,
        actor_id=actor.id, actor_label=actor.email, ip_address=ip,
        detail={"sync_ref": ref, "synced": len(synced_ids),
                "deposits": result["deposits"], "skipped": skipped,
                "files": result["files"], "backups": result["backups"],
                "retrain_required": False},
    )
    result["synced"] = len(synced_ids)
    result["message"] = (f"Synced {len(synced_ids)} ore observation(s) into "
                         f"{len(result['files'])} dataset file(s).")
    if job_id:
        jobs.finish(job_id, message=result["message"])
    return result


def invalidate_ml_caches() -> None:
    """Drop the pipeline's dataset caches so a sync takes effect immediately.

    `ore_loader` and `ore_grades` memoise their parsed files with
    `functools.lru_cache`, so without this a running process keeps serving the
    pre-sync ore map until it restarts — and the sync would look like it had
    done nothing.
    """
    import sys
    if str(REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(REPO_ROOT))
    cleared = []
    for mod_name in ("ml_pipeline.data_prep.ore_loader",
                     "ml_pipeline.data_prep.ore_grades"):
        try:
            mod = __import__(mod_name, fromlist=["*"])
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"could not reach {mod_name} to clear its cache: {exc}")
            continue
        for attr in dir(mod):
            fn = getattr(mod, attr)
            if hasattr(fn, "cache_clear"):
                fn.cache_clear()
                cleared.append(f"{mod_name}.{attr}")
    logger.info(f"cleared {len(cleared)} ml_pipeline dataset cache(s)")
    return None


# ═══════════════════════════════════════════════════════════════════════
# R11 — chemistry and groundwater levels
#
# Both writers follow the ore path exactly: resolve the approved rows, map them
# onto the file's real headers, back up, append with `record_source=added`, mark
# the observations synced, clear caches, audit. What differs is that these two
# files feed DERIVED artifacts — baselines and the flow field — so each result
# carries `stale_marks` naming what must now be recomputed before the change is
# visible to the engine. Appending alone is not enough for them, and saying so is
# the difference between a sync that works and one that only looks like it did.
# ═══════════════════════════════════════════════════════════════════════

async def _unsynced_water_samples(db: AsyncSession) -> list[dict[str, Any]]:
    """Approved chemistry, joined to the well that carries its coordinates."""
    rows = (await db.execute(text("""
        SELECT f.id::text AS obs_id,
               w.name AS location, w.latitude, w.longitude,
               d.name AS district,
               s.sampled_at, s.ph, s.ec_us_cm, s.carbonate_mg_l, s.bicarbonate_mg_l,
               s.chloride_mg_l, s.fluoride_mg_l, s.sulphate_mg_l, s.nitrate_mg_l,
               s.phosphate_mg_l, s.total_hardness, s.calcium_mg_l, s.magnesium_mg_l,
               s.sodium_mg_l, s.potassium_mg_l, s.iron_ppm, s.arsenic_ppb,
               s.uranium_ppb
        FROM field_observations f
        -- Join on applied_id, NOT target_id.
        --
        -- `ck_field_obs_target` enforces `operation = 'create' AND target_id IS
        -- NULL`, and every field-officer submission is a create — so target_id
        -- is ALWAYS null for exactly the rows this query exists to find. The id
        -- of the row approval created lands in `applied_id`. Joining on
        -- target_id matched nothing, so both syncs reported "Nothing to sync"
        -- while /dataset-sync/status correctly counted them as pending: the
        -- split-brain this whole feature exists to close.
        --
        -- COALESCE keeps the update path working, where target_id is the row
        -- being amended and applied_id is null.
        JOIN water_samples s   ON s.id = COALESCE(f.applied_id, f.target_id)
        JOIN monitoring_wells w ON w.id = s.well_id
        LEFT JOIN blocks b     ON b.id = w.block_id
        LEFT JOIN districts d  ON d.id = b.district_id
        WHERE f.observation_type = 'water_sample'
          AND f.status = 'approved'
          AND f.synced_to_dataset_at IS NULL
        ORDER BY f.reviewed_at
    """))).mappings().all()
    return [dict(r) for r in rows]


async def _unsynced_levels(db: AsyncSession) -> list[dict[str, Any]]:
    """Approved level readings, joined to their station."""
    rows = (await db.execute(text("""
        SELECT f.id::text AS obs_id,
               st.name AS station_name, st.latitude, st.longitude,
               d.name AS district,
               r.recorded_at, r.groundwater_level
        FROM field_observations f
        -- Join on applied_id, NOT target_id.
        --
        -- `ck_field_obs_target` enforces `operation = 'create' AND target_id IS
        -- NULL`, and every field-officer submission is a create — so target_id
        -- is ALWAYS null for exactly the rows this query exists to find. The id
        -- of the row approval created lands in `applied_id`. Joining on
        -- target_id matched nothing, so both syncs reported "Nothing to sync"
        -- while /dataset-sync/status correctly counted them as pending: the
        -- split-brain this whole feature exists to close.
        --
        -- COALESCE keeps the update path working, where target_id is the row
        -- being amended and applied_id is null.
        JOIN groundwater_level_readings r ON r.id = COALESCE(f.applied_id, f.target_id)
        JOIN monitoring_stations st       ON st.id = r.station_id
        LEFT JOIN blocks b                ON b.id = st.block_id
        LEFT JOIN districts d             ON d.id = b.district_id
        WHERE f.observation_type = 'groundwater_level'
          AND f.status = 'approved'
          AND f.synced_to_dataset_at IS NULL
        ORDER BY f.reviewed_at
    """))).mappings().all()
    return [dict(r) for r in rows]


#: DB column -> the header as it literally appears in waterQuality_jharkhand.csv.
#: Transcribed from the file, not guessed: "EC (µS/cm at" really is truncated
#: mid-unit in the source, and `S. No.` carries a UTF-8 BOM on the first column.
_WQ_MAP = {
    "ph": "pH", "ec_us_cm": "EC (µS/cm at", "carbonate_mg_l": "CO3 (mg/L)",
    "bicarbonate_mg_l": "HCO3", "chloride_mg_l": "Cl (mg/L)",
    "fluoride_mg_l": "F (mg/L)", "sulphate_mg_l": "SO4", "nitrate_mg_l": "NO3",
    "phosphate_mg_l": "PO4", "total_hardness": "Total Hardness",
    "calcium_mg_l": "Ca (mg/L)", "magnesium_mg_l": "Mg (mg/L)",
    "sodium_mg_l": "Na (mg/L)", "potassium_mg_l": "K (mg/L)",
    "iron_ppm": "Fe (ppm)", "arsenic_ppb": "As (ppb)", "uranium_ppb": "U (ppb)",
}


async def sync_water_quality(db: AsyncSession, *, actor, dry_run: bool = False,
                             ip: Optional[str] = None) -> dict[str, Any]:
    """Append approved chemistry to `waterQuality_jharkhand.csv`."""
    import pandas as pd

    f = dsx.get("water_quality")
    items = await _unsynced_water_samples(db)
    ref = uuid.uuid4().hex[:12]
    result: dict[str, Any] = {
        "sync_ref": ref, "dry_run": dry_run, "count": len(items),
        "files": [], "backups": [], "retrain_required": False,
        "stale_marks": list(f.stale_marks),
        "note": ("Chemistry rows set the ambient baselines an excursion UCL is "
                 "computed against. Appending them is not enough - run "
                 "POST /model-ops/recompute-baselines for the engine to use them."),
    }
    if not items:
        result["message"] = "Nothing to sync."
        return result

    job_id = None if dry_run else jobs.start(
        "sync_water_quality", label="Syncing approved chemistry", actor=getattr(actor, "email", None))

    df = dsx._read(f)
    next_no = int(pd.to_numeric(df["S. No."], errors="coerce").max() or 0) + 1

    new_rows = []
    for i, it in enumerate(items):
        row: dict[str, Any] = {
            "S. No.": next_no + i,
            "State": "Jharkhand",
            "District": it["district"] or "",
            "Location": it["location"] or "",
            "Longitude": round(float(it["longitude"]), 4),
            "Latitude": round(float(it["latitude"]), 4),
            "Year": it["sampled_at"].year if it["sampled_at"] else "",
            # `source_table` is the file's own provenance column (which CGWB
            # table a row came from). It is NOT record_source, and conflating
            # them would lose the distinction between "where this measurement
            # was published" and "did this system add it".
            "source_table": "field_observation",
            dsx.SOURCE_COL: dsx.SOURCE_ADDED,
            dsx.REF_COL: ref,
        }
        for db_col, csv_col in _WQ_MAP.items():
            v = it.get(db_col)
            row[csv_col] = "" if v is None else v
        new_rows.append(row)

    if dry_run:
        result["message"] = f"Would append {len(new_rows)} chemistry row(s)."
        result["preview"] = new_rows[:3]
        return result

    result["backups"].append(str(dsx.backup_file(f, ref).relative_to(REPO_ROOT)))
    # `append_rows`, not a pandas rewrite: round-tripping the frame reformatted
    # every untouched number in the file (0.00 -> 0.0), so appending three rows
    # produced a 794-line diff. See datasets.py "raw CSV access".
    dsx.append_rows(f, new_rows, ref)
    result["files"].append(f.relpath)

    await _mark_synced(db, [it["obs_id"] for it in items], ref)
    dsx.invalidate_caches()
    await audit.record(
        action="dataset.sync_water_quality", entity_type="datasets", entity_id=ref,
        actor_id=actor.id, actor_label=actor.email, ip_address=ip,
        detail={"sync_ref": ref, "synced": len(items), "files": result["files"],
                "backups": result["backups"], "stale_marks": result["stale_marks"]},
    )
    result["synced"] = len(items)
    result["message"] = (f"Synced {len(items)} chemistry observation(s). "
                         f"Baselines are now stale - recompute them.")
    if job_id:
        jobs.finish(job_id, message=result["message"])
    return result


async def sync_groundwater_levels(db: AsyncSession, *, actor, dry_run: bool = False,
                                  ip: Optional[str] = None) -> dict[str, Any]:
    """Append approved level readings to `cgwb_waterlevel_jharkhand.csv`."""
    import pandas as pd

    f = dsx.get("groundwater_levels")
    items = await _unsynced_levels(db)
    ref = uuid.uuid4().hex[:12]
    result: dict[str, Any] = {
        "sync_ref": ref, "dry_run": dry_run, "count": len(items),
        "files": [], "backups": [], "retrain_required": False,
        "stale_marks": list(f.stale_marks),
        "note": ("Level readings set the groundwater flow field - gradient and "
                 "azimuth at every pin. Appending them is not enough: run "
                 "POST /model-ops/rebuild-flow-field, which needs the GLO-30 DEM "
                 "on disk, for the engine to use them."),
    }
    if not items:
        result["message"] = "Nothing to sync."
        return result

    job_id = None if dry_run else jobs.start(
        "sync_groundwater_levels", label="Syncing approved level readings", actor=getattr(actor, "email", None))

    df = dsx._read(f)
    next_id = int(pd.to_numeric(df["id"], errors="coerce").max() or 0) + 1

    new_rows = []
    for i, it in enumerate(items):
        rec = it["recorded_at"]
        new_rows.append({
            "id": next_id + i,
            "date": rec.date().isoformat() if rec else "",
            "state_name": "Jharkhand", "state_code": 20,
            "district_name": it["district"] or "", "district_code": "",
            "station_name": it["station_name"] or "",
            "latitude": round(float(it["latitude"]), 5),
            "longitude": round(float(it["longitude"]), 5),
            "basin": "", "sub_basin": "",
            # The file's existing `source` column means the collecting agency
            # (CGWB). A field-submitted reading is not CGWB's, so it says so.
            "source": "FIELD",
            "currentlevel": float(it["groundwater_level"]),
            "level_diff": "",
            dsx.SOURCE_COL: dsx.SOURCE_ADDED,
            dsx.REF_COL: ref,
        })

    if dry_run:
        result["message"] = f"Would append {len(new_rows)} level reading(s)."
        result["preview"] = new_rows[:3]
        return result

    result["backups"].append(str(dsx.backup_file(f, ref).relative_to(REPO_ROOT)))
    dsx.append_rows(f, new_rows, ref)
    result["files"].append(f.relpath)

    await _mark_synced(db, [it["obs_id"] for it in items], ref)
    dsx.invalidate_caches()
    await audit.record(
        action="dataset.sync_groundwater_levels", entity_type="datasets",
        entity_id=ref, actor_id=actor.id, actor_label=actor.email, ip_address=ip,
        detail={"sync_ref": ref, "synced": len(items), "files": result["files"],
                "backups": result["backups"], "stale_marks": result["stale_marks"]},
    )
    result["synced"] = len(items)
    result["message"] = (f"Synced {len(items)} level reading(s). "
                         f"The flow field is now stale - rebuild it.")
    if job_id:
        jobs.finish(job_id, message=result["message"])
    return result


async def _mark_synced(db: AsyncSession, obs_ids: list[str], ref: str) -> None:
    if not obs_ids:
        return
    stmt = text("""
        UPDATE field_observations
        SET synced_to_dataset_at = now(), dataset_sync_ref = :ref
        WHERE id IN :ids
    """).bindparams(sa_bindparam("ids", expanding=True))
    await db.execute(stmt, {"ref": ref, "ids": [uuid.UUID(i) for i in obs_ids]})
    await db.commit()


async def sync_all(db: AsyncSession, *, actor, dry_run: bool = False,
                   ip: Optional[str] = None) -> dict[str, Any]:
    """Run every syncable type in one deliberate action."""
    parts = {
        "ore_presence": await sync_ore(db, actor=actor, dry_run=dry_run, ip=ip),
        "water_sample": await sync_water_quality(db, actor=actor, dry_run=dry_run, ip=ip),
        "groundwater_level": await sync_groundwater_levels(
            db, actor=actor, dry_run=dry_run, ip=ip),
    }
    stale: set[str] = set()
    for p in parts.values():
        stale.update(p.get("stale_marks", []))
    total = sum(p.get("synced", 0) for p in parts.values())
    return {
        "dry_run": dry_run, "synced": total, "by_type": parts,
        "stale_marks": sorted(stale),
        "message": (f"Synced {total} observation(s) across "
                    f"{sum(1 for p in parts.values() if p.get('synced'))} type(s)."
                    if total else "Nothing to sync."),
    }


async def reconcile_orphans(db: AsyncSession, *, actor, dry_run: bool = False,
                            ip: Optional[str] = None) -> dict[str, Any]:
    """Resolve approved observations whose applied row no longer exists.

    These can never sync: the sync joins to the row approval created, and that
    row is gone — so the endpoint reports "Nothing to sync" while the status
    keeps counting them as pending, for ever. A queue that cannot be emptied is
    worse than a full one, because it trains people to ignore the number.

    Marked `rejected` with a review note stating exactly why, rather than
    silently marked synced: they never reached the datasets, and recording that
    they did would be a lie told to the audit log. Re-submitting is the way to
    get the observation back.
    """
    rows = (await db.execute(text("""
        SELECT f.id::text, f.observation_type, f.submitted_by::text,
               f.applied_id::text, f.note
        FROM field_observations f
        WHERE f.status = 'approved' AND f.synced_to_dataset_at IS NULL
          AND CASE f.observation_type
                WHEN 'water_sample' THEN
                  NOT EXISTS (SELECT 1 FROM water_samples w
                              WHERE w.id = COALESCE(f.applied_id, f.target_id))
                WHEN 'groundwater_level' THEN
                  NOT EXISTS (SELECT 1 FROM groundwater_level_readings r
                              WHERE r.id = COALESCE(f.applied_id, f.target_id))
                ELSE false END
    """))).mappings().all()

    items = [dict(r) for r in rows]
    result: dict[str, Any] = {
        "dry_run": dry_run,
        "count": len(items),
        "items": items,
        "note": ("Marked rejected, not synced: these never reached the datasets, "
                 "and recording that they did would put a false statement in the "
                 "audit log. Re-submit if the observation is still wanted."),
    }
    if dry_run or not items:
        result["message"] = (
            f"{len(items)} observation(s) can no longer be synced."
            if items else "Nothing to reconcile — every approved observation is "
                          "either synced or still syncable.")
        return result

    # `applied_id` must be cleared alongside the status change: the constraint
    # `ck_field_obs_applied` is `applied_id IS NULL OR status = 'approved'`, so
    # rejecting while the pointer still dangles fails outright. Clearing it is
    # also the honest record — it points at a row that no longer exists.
    stmt = text("""
        UPDATE field_observations
        SET status = 'rejected',
            reviewed_at = now(),
            applied_id = NULL,
            review_note = :why
        WHERE id IN :ids
    """).bindparams(sa_bindparam("ids", expanding=True))
    await db.execute(stmt, {
        "why": ("Cannot be synced: the record this created was deleted, so there "
                "is nothing left to carry into the datasets. Re-submit if still "
                "needed."),
        "ids": [uuid.UUID(i["id"]) for i in items],
    })
    await db.commit()

    await audit.record(
        action="dataset.reconcile_orphans", entity_type="field_observations",
        entity_id="orphans", actor_id=actor.id, actor_label=actor.email,
        ip_address=ip,
        detail={"count": len(items),
                "ids": [i["id"] for i in items],
                "types": sorted({i["observation_type"] for i in items})},
    )
    result["message"] = (
        f"Resolved {len(items)} observation(s) that could never sync. The queue "
        f"now reflects only work that can actually be done.")
    return result
